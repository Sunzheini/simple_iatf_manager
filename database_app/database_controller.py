import sqlite3
from tools import sqlite3_code_generator
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class DatabaseController:
    def __init__(self, database_name):
        self.database_name = database_name

    @staticmethod
    def _open_db_connection(database_name):
        conn = sqlite3.connect(database_name)
        c = conn.cursor()
        return conn, c

    @staticmethod
    def _commit_and_close_db_connection(conn):
        conn.commit()
        conn.close()

    def create_database_and_table(self, table_name, table_structure):
        create_table_sqlite3_code = \
            sqlite3_code_generator.create_table_string(table_name, table_structure)
        conn, c = self._open_db_connection(self.database_name)
        c.execute(create_table_sqlite3_code)
        self._commit_and_close_db_connection(conn)
        print(f"table {table_name} created")
        return f"table {table_name} created"

    def empty_table(self):
        pass

    def drop_table(self, table_name):
        drop_table_sqlite3_code = sqlite3_code_generator.drop_table
        conn, c = self._open_db_connection(self.database_name)
        custom_table_code = drop_table_sqlite3_code + table_name
        try:
            c.execute(custom_table_code)
            self._commit_and_close_db_connection(conn)
            print(f"table {table_name} deleted")
            return f"table {table_name} deleted"
        except sqlite3.OperationalError:
            print("Invalid database name")
            return "Invalid database name"

    def load_from_excel(self, table_name):
        current_file_path = filedialog.askopenfilename()  # returns a string where the file is located

        # ToDo: not an excel file exception
        workbook = load_workbook(current_file_path)
        worksheet = workbook[workbook.sheetnames[0]]
        for row in range(16, 30):  # 1 is first row, not 0
            current_row = []
            for col in range(1, 12):  # 1 is first col, not 0 // 2, 3 e za 1 kolona samo
                char = get_column_letter(col)  # == chr(65 + col)
                current_row.append(worksheet[char + str(row)].value)

            insert_sqlite3_code = sqlite3_code_generator.insert_single_row(table_name, current_row)
            conn, c = self._open_db_connection(self.database_name)
            c.execute(insert_sqlite3_code)
            self._commit_and_close_db_connection(conn)

        text = f"Loaded from: {current_file_path}"
        print(text)
        return text
